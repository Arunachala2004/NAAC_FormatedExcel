import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

# Function to process the Excel file
def process_excel(input_df):
    # Combine author names
    input_df['Author Names'] = input_df[[
        'Author Name', 'Co Author 1', 'Co Author 2', 'Co Author 3', 'Co Author 4'
    ]].fillna('').apply(lambda x: '\n'.join(filter(None, x)), axis=1)  # Use \n for line breaks

    # Combine month and year into a single field
    input_df['Month and Year of Publication'] = pd.to_datetime(
        input_df['Month of publication'] + ' ' + input_df['Year of publication'].astype(str),
        errors='coerce'
    ).dt.strftime('%b %Y')  # Formatting as 'MMM YYYY'

    # Select and rename columns
    processed_df = input_df[[
        'Author Names', 'Department', 'Title of the Paper', 'Name of the journal',
        'Month and Year of Publication', 'ISSN Number', 'Link of the journal'
    ]].rename(columns={
        'Author Names': 'Name of the Author(s)',
        'Department': 'Department of the Author(s)',
        'Title of the Paper': 'Title of the Paper',
        'Name of the journal': 'Name of the Journal',
        'Month and Year of Publication': 'Month and Year of publication',
        'ISSN Number': 'ISSN',
        'Link of the journal': 'Link to the notification in UGC enlistment of the Journal'
    })

    return processed_df

# Function to style and save the final output as Excel
def style_and_save_excel(processed_df):
    # Save the processed data to Excel
    processed_file_path = 'processed_output.xlsx'
    processed_df.to_excel(processed_file_path, index=False, sheet_name='Processed Data')

    # Load the workbook and apply styles using openpyxl
    workbook = load_workbook(processed_file_path)
    sheet = workbook.active

    # Adjust column widths and row heights
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = 19

    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 100

    # Format the 'Name of the Author(s)' column with bold for the first author and text wrapping
    for row in range(2, sheet.max_row + 1):  # Skip header row
        cell = sheet.cell(row=row, column=1)  # Column 1 is 'Name of the Author(s)'
        author_names = cell.value.split('\n')
        primary_author = author_names[0]  # Primary author is the first name
        formatted_names = []
        for idx, author in enumerate(author_names):
            if idx == 0:
                # Bold the first author
                formatted_names.append(f"{author}")
            else:
                formatted_names.append(author)

        # Set the cell value with each author's name on a new line
        cell.value = "\n".join(formatted_names)  # Store each name on a new line
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Bold only the primary author in the first line
        cell.font = Font(bold=True if cell.value.split("\n")[0] == primary_author else False)

    # Apply center alignment to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Save the workbook to a BytesIO object for streaming
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# Streamlit UI
st.title("Excel Data(RAW) to NAAC formatted excel")
st.write("Upload your Excel file below and process it.")

uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx'])

if uploaded_file:
    try:
        # Read the uploaded file
        input_data = pd.ExcelFile(uploaded_file)
        input_df = input_data.parse('Form Responses 1')

        # Process the data
        processed_df = process_excel(input_df)

        # Display the processed dataframe
        st.subheader("Processed Data Preview")
        st.dataframe(processed_df)

        # Download the processed file
        st.subheader("Download Processed File")
        processed_output = style_and_save_excel(processed_df)
        st.download_button(
            label="Download Processed Excel",
            data=processed_output,
            file_name="processed_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"An error occurred: {e}")
